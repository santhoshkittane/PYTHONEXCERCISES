import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def create_excel_example(filename="company_sales.xlsx"):
    # 1. Initialize a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Q1 Sales"  # Rename the default sheet title

    # 2. Add header data
    headers = ["Product", "Units Sold", "Price Per Unit", "Total Revenue"]
    ws.append(headers)

    # 3. Add row data
    data = [
        ["Laptops", 15, 1200],
        ["Monitors", 45, 300],
        ["Keyboards", 120, 50],
        ["Mice", 150, 25]
    ]
    for row in data:
        ws.append(row)

    # 4. Insert Excel formulas dynamically
    # Calculate Total Revenue (Units Sold * Price) -> Column B * Column C
    for row_idx in range(2, 6):  # Rows 2 to 5
        ws[f"D{row_idx}"] = f"=B{row_idx}*C{row_idx}"

    # Add a Totals Row
    ws["A6"] = "Totally"
    ws["B6"] = "=SUM(B2:B5)"
    ws["D6"] = "=SUM(D2:D5)"

    # 5. Apply Styles and Formatting
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)

    # Style the headers (Row 1)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Style the summary rows (Row 6)
    ws["A6"].font = bold_font
    ws["B6"].font = bold_font
    ws["D6"].font = bold_font

    # 6. Save the spreadsheet
    wb.save(filename)
    print(f"Workbook successfully saved as '{filename}'\n")


def read_excel_example(filename="company_sales.xlsx"):
    print("--- Reading data from Excel file ---")
    
    # Load the existing workbook
    wb = openpyxl.load_workbook(filename)
    ws = wb["Q1 Sales"]

    # Iterate through rows and columns using iter_rows()
    # values_only=False allows us to see formulas rather than values
    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=4):
        row_values = [cell.value for cell in row]
        print(row_values)


# Execute the program
if __name__ == "__main__":
    create_excel_example()
    read_excel_example()
