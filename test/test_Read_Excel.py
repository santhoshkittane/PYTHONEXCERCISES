from openpyxl import load_workbook

#added

def test_read_excel_file(file_path):
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook.active
    print(f"Reading Sheet: {sheet.title}\n")
    for row in sheet.iter_rows(values_only=True):
        print(row)
        if row[0] == 'UserMail':
            continue
        user_mail, user_name, password = row
        print(f"Email: {user_mail}")
        print(f"Username: {user_name}")
        print(f"Password: {password}")
        print("-" * 20)
    

if __name__ == "__main__":
    excel_file = input("Enter the path to the Excel file (e.g., C:\\path\\to\\file.xlsx): ").strip()
    test_read_excel_file(excel_file)
    # read_excel_file("C:\\Users\\SanthoshKittane\\PycharmProjects\\SimpleCalc\\my_package.egg-info\\SRC\\Test.xlsx")